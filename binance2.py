from openpyxl import load_workbook    
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


driver = webdriver.Chrome("C:\\Users\\Batu\\Desktop\\chromedriver.exe")
url = "https://www.binance.com/tr/markets" #url of the site
driver.get(url)

# variables which will store names, prices, price changes and market values
names = [] 
prices = [] 
changes = [] 
market_values = []

wb = load_workbook("C:\\Users\\Batu\\Desktop\\Crypto Currencies.xlsx")# the excel file
ws = wb.active
ws.title = "Crypto Currencies" # adjustes the title

if(ws["A1"].value != ""):
    ws.delete_rows(0,(ws.max_row + 1)) # clears the excel sheet

#setting the titles
ws["A1"] = "Name"
ws["B1"] = "Price"
ws["C1"] = "Change" 
ws["D1"] = "M Value"

for col in range(1,5):
    ws[get_column_letter(col) + "1"].font = Font(bold = True) # makes the headlines bold

for i in range(8):
    driver.implicitly_wait(2)
    coin_names = driver.find_elements(By.CLASS_NAME, "css-1ap5wc6") # finds coin names
    coin_prices = driver.find_elements(By.XPATH, "//*[@class ='css-ovtrou' or @class = 'css-li1e6c' or @class = 'css-1ez6tx0']") # finds coin prices
    price_changes = driver.find_elements(By.XPATH, "//*[@class='css-1ca67uc' or @class=  'css-1vgqjs4']") #finds price changes
    values = driver.find_elements(By.CLASS_NAME, "css-s779xv") # finds market values   
    driver.implicitly_wait(2)

    #appending the data collected from site to relevant lists
    for name in coin_names:
        names.append(name.text)
    for price in coin_prices:
        prices.append(price.text)
    for value in values:
        market_values.append(value.text) 
    for change in price_changes:
        changes.append(change.text) 

    # writes the necessary info to excel row by row
    for row in range(2,len(names)+2):
        ws.append([names[row-2],prices[row-2],changes[row-2],market_values[row-2]]) 
  
    #clears all lists for going to next page
    names.clear() 
    prices.clear()
    changes.clear()
    market_values.clear()

    # goes to next page if it exists
    if(i != 8): 
        next_page_button = driver.find_element(By.XPATH, "//*[@aria-label ='Next page']")

        driver.implicitly_wait(2)
        action = webdriver.ActionChains(driver)
        action.move_to_element(next_page_button)
        driver.implicitly_wait(2)
        action.click(next_page_button)
        action.perform()


wb.save("C:\\Users\\Batu\\Desktop\\Crypto Currencies.xlsx") # saves the document
driver.implicitly_wait(5)
driver.close()




